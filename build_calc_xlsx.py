# build_calc_xlsx.py
# -----------------------------------------------------------------------------
# Campaign Technology Suite - monthly pricing (live Excel calculator), 2 sheets:
#   Sheet 1 "Pricing"        -> per-tool: Salary (A) + Tool Cost (B) = Grand Total
#   Sheet 2 "Tool Cost (B)"  -> full breakdown of B (AI, server, proxy, etc.)
# Master Tool-Cost cell is LINKED to the breakdown total, so everything is live.
# -----------------------------------------------------------------------------
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY="0E2A5E"; NAVY2="1B3F86"; SAFFRON="FF7A00"; GREEN="2E9E5B"
LIGHT="F3F6FC"; LIGHT2="E8EEF8"; INK="1A2433"; GREY="52607A"; WHITE="FFFFFF"
INR = '"₹" #,##,##0'

thin = Side(style="thin", color="C9D4E8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def styler(ws):
    def cell(r, c, val, *, bold=False, size=10, color=INK, fill=None, align="left",
             numfmt=None, white=False, wrap=False):
        cc = ws.cell(row=r, column=c, value=val)
        cc.font = Font(name="Calibri", size=size, bold=bold,
                       color=(WHITE if white else color))
        if fill: cc.fill = PatternFill("solid", fgColor=fill)
        cc.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if numfmt: cc.number_format = numfmt
        cc.border = border
        return cc
    return cell

# =============================================================================
# SHEET 2 first: Tool Cost (B) breakdown  -> total cell referenced by Sheet 1
# =============================================================================
wsb = wb.active; wsb.title = "Tool Cost (B)"
wsb.sheet_view.showGridLines = False
wsb.column_dimensions["A"].width = 30
wsb.column_dimensions["B"].width = 60
wsb.column_dimensions["C"].width = 18
cb = styler(wsb)

wsb.merge_cells("A1:C1")
cb(1, 1, "Narrative Intelligence  -  Tool Cost / Month (B) Breakdown", bold=True,
   size=15, fill=NAVY, white=True)
wsb.row_dimensions[1].height = 28
wsb.merge_cells("A2:C2")
cb(2, 1, "What the monthly tool / infrastructure cost covers (excludes team salaries)",
   size=10, fill=NAVY, white=True)
wsb.row_dimensions[2].height = 18
wsb.row_dimensions[3].height = 6

bhr = 4
cb(bhr, 1, "Component", bold=True, fill=NAVY, white=True)
cb(bhr, 2, "What it covers", bold=True, fill=NAVY, white=True)
cb(bhr, 3, "Monthly (INR)", bold=True, fill=NAVY, white=True, align="right")
wsb.row_dimensions[bhr].height = 22

b_items = [
    ("Claude AI / Vision Analysis",
     "~10,000 ads/day read & classified by Claude AI + narrative forecast & "
     "counter-strategy (Anthropic API billing)", 150000),
    ("Server & Cloud Infrastructure",
     "24/7 production server, compute, uptime, load handling", 60000),
    ("Proxy & IP Rotation",
     "Rotating proxies / IPs for reliable large-scale data collection", 40000),
    ("Database, Storage & Archive",
     "Permanent record of 3,00,000+ ads/month, indexing & backups", 30000),
    ("Maintenance, DevOps & Security",
     "Updates, monitoring, security patches, incident response", 30000),
    ("Domain, SSL, API & Misc",
     "Domain, SSL, Meta Ad Library API access & misc infrastructure", 10000),
]
bf = bhr + 1
r = bf
for i, (comp, desc, amt) in enumerate(b_items):
    fill = WHITE if i % 2 == 0 else LIGHT
    cb(r, 1, comp, bold=True, fill=fill)
    cb(r, 2, desc, size=9, color=GREY, fill=fill, wrap=True)
    cb(r, 3, amt, bold=True, fill=fill, align="right", numfmt=INR)
    wsb.row_dimensions[r].height = 30
    r += 1
bl = r - 1
B_TOTAL_ROW = r
cb(r, 1, "TOTAL  -  Tool Cost / Month (B)", bold=True, size=11, white=True, fill=SAFFRON)
cb(r, 2, "", fill=SAFFRON)
cb(r, 3, f"=SUM(C{bf}:C{bl})", bold=True, size=12, white=True, fill=SAFFRON,
   align="right", numfmt=INR)
wsb.row_dimensions[r].height = 24
wsb.freeze_panes = "A5"

# =============================================================================
# SHEET 1: Pricing master  (Salary A + Tool Cost B = Grand Total)
# =============================================================================
ws = wb.create_sheet("Pricing", 0)   # make it the first sheet
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 34
for col in ("B", "C", "D", "E", "F"):
    ws.column_dimensions[col].width = 17
cp = styler(ws)

ws.merge_cells("A1:F1")
cp(1, 1, "Campaign Technology Suite  -  Monthly Pricing", bold=True, size=16,
   fill=NAVY, white=True)
ws.row_dimensions[1].height = 30
ws.merge_cells("A2:F2")
cp(2, 1, "AAP Punjab 2027   |   Salary Total (A) + Tool Cost/Month (B) = Grand Total"
   "   |   Tool Cost breakdown on the 'Tool Cost (B)' sheet",
   size=10, fill=NAVY, white=True)
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 6

hr = 4
heads = ["Tool Name / Description", "Employees", "Avg Salary (₹)",
         "Salary Total (A)", "Tool Cost / Month (B)", "Grand Total (A+B)"]
for i, h in enumerate(heads, start=1):
    cp(hr, i, h, bold=True, fill=NAVY, white=True,
       align=("left" if i == 1 else "center"), wrap=True)
ws.row_dimensions[hr].height = 30

# Narrative Intelligence: B is LINKED to the breakdown total on Sheet 2.
first = hr + 1
r = first
# NI row
cp(r, 1, "Narrative Intelligence", bold=True, fill=WHITE)
cp(r, 2, 6, fill=WHITE, align="center")
cp(r, 3, 60000, fill=WHITE, align="right", numfmt=INR)
cp(r, 4, f"=B{r}*C{r}", fill=WHITE, align="right", numfmt=INR)               # A
cp(r, 5, f"='Tool Cost (B)'!C{B_TOTAL_ROW}", fill=WHITE, align="right",
   numfmt=INR)                                                              # B (linked)
cp(r, 6, f"=D{r}+E{r}", bold=True, color=NAVY, fill=WHITE, align="right",
   numfmt=INR)                                                              # A+B
ws.row_dimensions[r].height = 22
r += 1
# blank tool rows (formulas ready)
for i in range(5):
    fill = LIGHT if i % 2 == 0 else WHITE
    cp(r, 1, None, fill=fill)
    cp(r, 2, None, fill=fill, align="center")
    cp(r, 3, None, fill=fill, align="right", numfmt=INR)
    cp(r, 4, f"=B{r}*C{r}", fill=fill, align="right", numfmt=INR)
    cp(r, 5, None, fill=fill, align="right", numfmt=INR)
    cp(r, 6, f"=D{r}+E{r}", bold=True, color=NAVY, fill=fill, align="right", numfmt=INR)
    ws.row_dimensions[r].height = 22
    r += 1
last = r - 1

cp(r, 1, "TOTAL  (all tools)", bold=True, size=11, white=True, fill=SAFFRON)
cp(r, 2, "", fill=SAFFRON); cp(r, 3, "", fill=SAFFRON)
cp(r, 4, f"=SUM(D{first}:D{last})", bold=True, white=True, fill=SAFFRON, align="right", numfmt=INR)
cp(r, 5, f"=SUM(E{first}:E{last})", bold=True, white=True, fill=SAFFRON, align="right", numfmt=INR)
cp(r, 6, f"=SUM(F{first}:F{last})", bold=True, size=12, white=True, fill=SAFFRON, align="right", numfmt=INR)
ws.row_dimensions[r].height = 26
note = r + 2
ws.merge_cells(f"A{note}:F{note}")
cp(note, 1, "Note: editable & live. Tool Cost (B) auto-pulls from the 'Tool Cost (B)' "
   "sheet. Change any value there or here and Salary Total, Grand Total and the suite "
   "TOTAL recompute automatically. Add more tools in the blank rows.",
   size=8.5, color=GREY, wrap=True)
ws.row_dimensions[note].height = 30
ws.freeze_panes = "A5"

out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Narrative_Intelligence_Calculation.xlsx")
wb.save(out)
print("WROTE", out)
