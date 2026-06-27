# app.py
# -----------------------------------------------------------------------------
# Flask app: routes, session-based login, APScheduler (6 ghante refresh +
# in-memory cache), aur Meta Ad Library data serve karta hai.
#
# Run locally:  python app.py
# Production:   gunicorn app:app   (Railway Procfile)
# -----------------------------------------------------------------------------

import os
import logging
import threading
from datetime import datetime, timezone, timedelta
from functools import wraps

# IMPORTANT: .env ko config import se PEHLE load karo, taaki local dev mein
# env vars config.py tak pahunch jaaye. Railway pe env already set hote hain.
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

from io import BytesIO

from flask import (Flask, jsonify, redirect, render_template, request,
                   session, url_for, send_file)
from apscheduler.schedulers.background import BackgroundScheduler

import config
import meta_api
import classifier
import strategy
import history
import intelligence
import archive

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
log = logging.getLogger("app")

app = Flask(__name__)
app.secret_key = config.SECRET_KEY
# Login ko lambe time tak yaad rakho — browser band/reopen ya server restart ke
# baad bhi dobara id/password na maange (default 30 din, env se badal sakte ho).
app.permanent_session_lifetime = timedelta(
    days=int(os.environ.get("SESSION_DAYS", "30")))
app.config.update(
    SESSION_REFRESH_EACH_REQUEST=True,   # har visit pe expiry aage badh jaaye
    SESSION_COOKIE_SAMESITE="Lax",
)


@app.after_request
def _no_cache_html(resp):
    """HTML pages ko cache mat karo — browser hamesha latest dashboard le.
    (Isse 'purana cached version atak gaya' wali dikkat khatam hoti hai.)"""
    ctype = resp.headers.get("Content-Type", "")
    if ctype.startswith("text/html"):
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
    return resp

# =============================================================================
# In-memory cache — scheduler isse har 6 ghante refresh karta hai.
# =============================================================================
CACHE = {
    "mode": "demo",
    "count": 0,
    "ads": [],
    "party_spend": [],
    "top_spender": "N/A",
    "top_themes": [],
    "errors": [],
    "updated_at": None,
    # ready=False -> abhi pehli fetch chal rahi hai (warm-up). Frontend isse
    # dekh ke "loading" dikhata hai aur auto-poll karta hai jab tak ready na ho.
    "ready": False,
}
_CACHE_LOCK = threading.Lock()

# =============================================================================
# Tracking ON/OFF — dashboard se start/stop. Band hone pe koi Meta fetch / Claude
# classify nahi hota (cost bachta hai). State file mein save hota hai taaki
# restart ke baad bhi yaad rahe.
# =============================================================================
_STATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
_STATE_FILE = os.path.join(_STATE_DIR, "state.json")


def _load_tracking():
    try:
        import json
        with open(_STATE_FILE) as f:
            return bool(json.load(f).get("tracking", True))
    except Exception:
        return True  # default: tracking ON


def _save_tracking(on):
    try:
        import json
        os.makedirs(_STATE_DIR, exist_ok=True)
        with open(_STATE_FILE, "w") as f:
            json.dump({"tracking": bool(on)}, f)
    except Exception as e:
        log.warning("save tracking state failed: %s", e)


TRACKING_ENABLED = _load_tracking()


def _payload_from_archive(errors):
    """
    Live Meta data na ho to archive ke (pehle-se-classified) real ads se ek
    dashboard payload banao — demo se behtar. None agar archive khali ho.
    """
    try:
        ads = archive.dashboard_ads(config.MAX_TOTAL_ADS)
    except Exception as e:
        log.warning("archive fallback load failed: %s", e)
        return None
    if not ads:
        return None
    payload = {"mode": "archive", "count": len(ads), "ads": ads,
               "errors": errors or []}
    # UI notice — kyun saved data dikha rahe hain (token expire / Meta issue).
    # Warm-up seed (errors khaali) pe koi notice nahi.
    joined = " ".join(errors or []).lower()
    if any(k in joined for k in ("expire", "validating access token",
                                 "session has expired", "access token")):
        payload["notice"] = ("Token expire/invalid ho gaya — purana saved data "
                             "dikhaya ja raha hai (sab data safe).")
    elif "permission" in joined:
        payload["notice"] = ("Meta token ko Ad Library permission nahi — saved "
                             "data dikhaya ja raha hai (sab data safe).")
    elif errors:
        payload["notice"] = ("Meta abhi unavailable — purana saved data dikhaya "
                             "ja raha hai (sab data safe).")
    payload.update(meta_api._build_aggregates(ads))
    payload["official_count"] = sum(1 for a in ads if a.get("is_official"))
    payload["non_official_count"] = sum(1 for a in ads if not a.get("is_official"))
    payload.update(classifier.build_ai_aggregates(ads))
    try:
        meta_api.assign_damage_levels(ads)
    except Exception:
        pass
    return payload


def refresh_cache():
    """Meta se fresh data laao aur CACHE update karo (thread-safe)."""
    if not TRACKING_ENABLED:
        log.info("Tracking band hai — refresh skip (na Meta fetch, na AI cost).")
        return
    log.info("Refreshing ad cache from Meta Ad Library...")
    try:
        payload = meta_api.fetch_all_ads()
    except Exception as e:
        # Kuch bhi unexpected fail ho to app crash na ho — purana cache rakho.
        log.exception("refresh_cache failed: %s", e)
        return

    if payload.get("mode") == "live":
        # Claude se stance (against/support AAP) + party + narrative classify.
        # Sirf naye ads classify hote hain (cached), isliye yeh fast rehta hai.
        try:
            classifier.enrich_ads(payload.get("ads", []))
            # Pro-AAP ad opponent party ki nahi ho sakti — proxy party fix karo.
            meta_api.correct_proxy_party(payload.get("ads", []))
            payload.update(classifier.build_ai_aggregates(payload.get("ads", [])))
            # Damage Radar: stance ke baad anti-AAP ads ko threat-rank karo.
            meta_api.assign_damage_levels(payload.get("ads", []))
        except Exception as e:
            log.warning("AI classification skipped: %s", e)
    else:
        # Live Meta fail/throttle (ya warm-up) — demo ki jagah archive ke REAL,
        # pehle-se-classified ads dikhao. War-room hamesha asli data dikhe.
        arch = _payload_from_archive(payload.get("errors", []))
        if arch:
            payload = arch
            log.info("Live unavailable — serving %d ads from archive.",
                     payload["count"])

    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    payload["ready"] = True  # pehli fetch complete -> frontend ko data dikhao
    with _CACHE_LOCK:
        CACHE.update(payload)

    # Trend Over Time: is refresh ka snapshot history mein save karo.
    try:
        history.record_snapshot(payload)
    except Exception as e:
        log.warning("history record skipped: %s", e)

    # Ad Archive: har ad ka permanent record (active/stopped tracking).
    try:
        archive.record_ads(payload.get("ads", []), payload.get("mode", "demo"))
    except Exception as e:
        log.warning("archive record skipped: %s", e)

    # Fresh tracking: sirf is month + abhi chal rahi ads rakho — purani auto-delete.
    if payload.get("mode") == "live":
        try:
            archive.purge_old()
        except Exception as e:
            log.warning("purge_old skipped: %s", e)
    log.info("Cache updated: mode=%s count=%s top_spender=%s",
             payload.get("mode"), payload.get("count"),
             payload.get("top_spender"))


# =============================================================================
# Auth helpers
# =============================================================================
def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        # Desktop app (local) mein login skip; web pe normal auth.
        if config.DESKTOP_MODE or session.get("logged_in"):
            return view(*args, **kwargs)
        return redirect(url_for("login", next=request.path))
    return wrapped


# =============================================================================
# Routes
# =============================================================================
@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        user = request.form.get("username", "")
        pw = request.form.get("password", "")
        if user == config.ADMIN_USER and pw == config.ADMIN_PASS:
            session.permanent = True   # cookie ko 30 din tak zinda rakho
            session["logged_in"] = True
            session["user"] = user
            nxt = request.args.get("next") or url_for("dashboard")
            return redirect(nxt)
        error = "Galat username ya password."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    return render_template("index.html",
                           refresh_hours=config.REFRESH_HOURS,
                           user=session.get("user", "admin"))


@app.route("/api/ads")
@login_required
def api_ads():
    """Cached JSON return karta hai (fast — per-request live call nahi)."""
    with _CACHE_LOCK:
        # Shallow copy taaki response ke time lock chhota rahe.
        data = dict(CACHE)
    return jsonify(data)


@app.route("/api/refresh", methods=["POST"])
@login_required
def api_refresh():
    """Manual refresh trigger — abhi Meta se data laata hai."""
    refresh_cache()
    with _CACHE_LOCK:
        data = dict(CACHE)
    return jsonify({"ok": True, "mode": data.get("mode"),
                    "count": data.get("count"),
                    "updated_at": data.get("updated_at"),
                    "tracking": TRACKING_ENABLED})


@app.route("/api/tracking", methods=["GET", "POST"])
@login_required
def api_tracking():
    """Tracking ON/OFF. POST {on:true/false} -> set; GET -> current state.
    OFF hone pe scheduler/refresh kuch fetch ya classify nahi karega."""
    global TRACKING_ENABLED
    if request.method == "POST":
        if request.is_json:
            on = bool((request.get_json(silent=True) or {}).get("on"))
        else:
            on = request.form.get("on") in ("1", "true", "on", "yes")
        TRACKING_ENABLED = on
        _save_tracking(on)
        log.info("Tracking %s by admin.", "ON" if on else "OFF")
        if on:
            # Resume -> turant ek fresh fetch background mein chala do.
            threading.Thread(target=refresh_cache, daemon=True).start()
    return jsonify({"tracking": TRACKING_ENABLED})


@app.route("/api/translate", methods=["POST"])
@login_required
def api_translate():
    """Ek ad ka text Hindi mein translate karke return karta hai (cached)."""
    data = request.get_json(silent=True) or {}
    ad_id = data.get("id", "")
    text = data.get("text", "")
    if not text:
        return jsonify({"ok": False, "error": "no text"}), 400
    hindi = classifier.translate_to_hindi(ad_id, text)
    if hindi is None:
        return jsonify({"ok": False, "error": "AI off ya translate fail"})
    return jsonify({"ok": True, "hindi": hindi})


@app.route("/api/strategy", methods=["GET", "POST"])
@login_required
def api_strategy():
    """
    GET  -> last cached strategy brief (agar generate ho chuki ho).
    POST -> current ads se nayi strategy brief generate karo (Claude Opus).
    """
    if request.method == "POST":
        with _CACHE_LOCK:
            ads = list(CACHE.get("ads", []))
        result = strategy.generate_brief(ads)
        if result.get("generated"):
            result["generated_at"] = datetime.now(timezone.utc).isoformat()
        return jsonify(result)
    # GET -> cached
    return jsonify(strategy.get_cached())


@app.route("/api/counter", methods=["POST"])
@login_required
def api_counter():
    """Ek anti-AAP ad ka specific counter generate karo (Damage Radar)."""
    data = request.get_json(silent=True) or {}
    ad_id = data.get("id", "")
    text = data.get("text", "")
    narrative = data.get("narrative", "")
    audience = data.get("audience", "")
    if not text:
        return jsonify({"ok": False, "error": "no text"}), 400
    counter = classifier.generate_counter(ad_id, text, narrative, audience)
    if counter is None:
        return jsonify({"ok": False, "error": "AI off ya fail"})
    return jsonify({"ok": True, "counter": counter})


@app.route("/api/audience")
@login_required
def api_audience():
    """Audience Vulnerability — opponents kis segment ko target kar rahe."""
    with _CACHE_LOCK:
        ads = list(CACHE.get("ads", []))
    return jsonify(intelligence.audience_vulnerability(ads))


@app.route("/api/forecast", methods=["POST"])
@login_required
def api_forecast():
    """Narrative Forecast — Claude predict kare konsa narrative chadega."""
    with _CACHE_LOCK:
        ads = list(CACHE.get("ads", []))
    hist = history.load_history()
    return jsonify(intelligence.generate_forecast(hist, ads))


@app.route("/api/creative", methods=["POST"])
@login_required
def api_creative():
    """Counter-Ad Creative — ek narrative ka ready-to-post ad."""
    data = request.get_json(silent=True) or {}
    narrative = data.get("narrative", "")
    audience = data.get("audience", "")
    attack = data.get("attack_text", "")
    if not narrative:
        return jsonify({"ok": False, "error": "narrative chahiye"}), 400
    creative = classifier.generate_creative(narrative, audience, attack)
    if creative is None:
        return jsonify({"ok": False, "error": "AI off ya fail"})
    return jsonify({"ok": True, "creative": creative})


@app.route("/api/archive")
@login_required
def api_archive():
    """Ad archive — saari pool ki gayi ads (active + stopped) with dates."""
    status = request.args.get("status", "all")
    days = request.args.get("days")
    party = request.args.get("party", "ALL")
    stance = request.args.get("stance", "ALL")
    ads = archive.get_archive(status=status, days=days, party=party, stance=stance)
    return jsonify({"count": len(ads), "ads": ads, "stats": archive.stats()})


@app.route("/api/archive/stats")
@login_required
def api_archive_stats():
    return jsonify(archive.stats())


@app.route("/api/spend")
@login_required
def api_spend():
    """Spend Tracker — AAP ke khilaf total + party-wise + daily spend (archive se)."""
    return jsonify(archive.spend_tracker())


@app.route("/api/usage")
@login_required
def api_usage():
    """Super-admin billing — Claude (AI) ka token usage + estimated cost (INR).
    Extra gate: SUPER_ADMIN_PASS sahi ho tab hi data, warna locked."""
    key = request.args.get("key", "")
    if config.SUPER_ADMIN_PASS and key != config.SUPER_ADMIN_PASS:
        return jsonify({"available": False, "locked": True})
    import usage
    return jsonify(usage.summary())


@app.route("/api/directory")
@login_required
def api_directory():
    """Advertiser directory — har 'Paid for by' disclaimer ki ads/spend/stance."""
    return jsonify(archive.directory())


@app.route("/api/export/archive")
@login_required
def api_export_archive():
    """Poore archive ka Excel (active + stopped, with first/last seen)."""
    status = request.args.get("status", "all")
    days = request.args.get("days")
    ads = archive.get_archive(status=status, days=days, limit=100000)
    return _build_ads_xlsx(ads, "narrative_intelligence_archive.xlsx",
                           include_archive_cols=True)


@app.route("/api/history")
@login_required
def api_history():
    """Trend Over Time data — snapshots + auto-insights."""
    items = history.load_history()
    return jsonify({
        "count": len(items),
        "snapshots": items,
        "insights": history.compute_insights(items),
    })


def _build_ads_xlsx(ads, fname, include_archive_cols=False):
    """Ads list se ek styled .xlsx banao aur download response do.
    include_archive_cols=True -> Status/First Seen/Last Seen/Stopped At bhi."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        return jsonify({"ok": False, "error": "openpyxl not installed"}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ads"

    headers = [
        "Facebook Page", "Facebook Page URL", "Page ID", "Handle", "Party",
        "Source", "Stance", "Damage Level", "Narrative", "AI Summary",
        "Spend (range)", "Impressions (range)", "Audience", "Regions",
        "Platforms", "Started", "Ended", "Theme (keyword)", "Ad Text",
        "View Ad on Meta (link)", "Ad ID",
    ]
    widths = [26, 40, 16, 16, 8, 14, 12, 12, 22, 34, 20, 20, 22, 26, 18, 12, 14, 18, 60, 40, 18]
    if include_archive_cols:
        headers += ["Status", "First Seen", "Last Seen", "Stopped At"]
        widths += [12, 17, 17, 17]
    ws.append(headers)

    stance_word = {"against": "Against AAP", "support": "Pro-AAP",
                   "neutral": "Neutral", "unknown": ""}

    def _ts(t):
        return (t or "")[:16].replace("T", " ")

    for a in ads:
        aud = a.get("audience") or {}
        aud_str = (f"{aud.get('gender_pct','')}% {aud.get('gender_top','')}, "
                   f"{aud.get('age_top','')}") if aud else ""
        pid = str(a.get("page_id", "") or "").strip()
        page_url = f"https://www.facebook.com/{pid}" if pid else ""
        row = [
            a.get("page", ""), page_url, pid, a.get("handle", ""),
            a.get("party", ""), a.get("source", ""),
            stance_word.get(a.get("stance", ""), a.get("stance", "")),
            (a.get("damage_level") or "").upper(),
            a.get("narrative", "") or a.get("theme", ""),
            a.get("narrative_summary", ""), a.get("spend", ""), a.get("impr", ""),
            aud_str, ", ".join(a.get("regions", []) or []),
            ", ".join(a.get("plat", []) or []),
            a.get("start", "") or a.get("started", ""),
            a.get("stop", "") or "Running",
            a.get("theme", ""), a.get("text", ""), a.get("snapshot_url", ""),
            a.get("id", ""),
        ]
        if include_archive_cols:
            row += [
                "ACTIVE" if a.get("active") else "STOPPED",
                _ts(a.get("first_seen")), _ts(a.get("last_seen")),
                _ts(a.get("stopped_at")),
            ]
        ws.append(row)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F2A37")
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/export")
@login_required
def api_export():
    """Current (active) ads ka Excel export."""
    with _CACHE_LOCK:
        ads = list(CACHE.get("ads", []))
    return _build_ads_xlsx(ads, "narrative_intelligence_ads.xlsx")


@app.route("/health")
def health():
    """Railway/uptime checks ke liye simple health endpoint (no auth)."""
    return jsonify({"status": "ok", "mode": CACHE.get("mode"),
                    "count": CACHE.get("count")})


# =============================================================================
# Scheduler setup — har REFRESH_HOURS ghante cache refresh.
# =============================================================================
scheduler = BackgroundScheduler(daemon=True)


def start_scheduler():
    if scheduler.running:
        return
    scheduler.add_job(refresh_cache, "interval",
                      hours=config.REFRESH_HOURS,
                      id="refresh_ads", replace_existing=True)
    scheduler.start()
    log.info("Scheduler started — refresh every %s hours.",
             config.REFRESH_HOURS)


def _seed_cache_from_archive():
    """
    Boot pe TURANT archive ke real (pehle-se-classified) ads cache mein daal do,
    taaki dashboard kabhi demo/khali na dikhe jab tak slow live refresh chal raha
    ho. Live refresh complete hone pe yeh fresh data se replace ho jaata hai.
    """
    try:
        arch = _payload_from_archive([])
        if arch:
            arch["updated_at"] = datetime.now(timezone.utc).isoformat()
            arch["ready"] = True
            with _CACHE_LOCK:
                CACHE.update(arch)
            log.info("Cache seeded from archive: %d real ads.", arch["count"])
        # CRITICAL: classifier _CACHE ko archive ke good stances se prime karo.
        # Warna restart ke baad refresh poori 10k+ ads dobara classify karta hai
        # (rate-limit -> 'unknown' -> archive corrupt). Isse sirf naye/unknown
        # ads classify honge, classified ads safe rahenge.
        try:
            classifier.prime_cache(archive.get_archive(status="all", limit=50000))
        except Exception as e:
            log.warning("prime classifier cache failed: %s", e)
    except Exception as e:
        log.warning("seed from archive failed: %s", e)


# App startup pe ek baar warm-up fetch + scheduler chalao.
# gunicorn ke under bhi yeh import-time pe chalega.
def _bootstrap():
    _seed_cache_from_archive()   # instant real data (demo se bacho)
    refresh_cache()              # phir live Meta se update (jab data mile)
    start_scheduler()


# Background thread se warm-up (taaki gunicorn boot block na ho).
threading.Thread(target=_bootstrap, daemon=True).start()


if __name__ == "__main__":
    # Local dev server. Railway pe gunicorn use hota hai (Procfile).
    app.run(host="0.0.0.0", port=config.PORT, debug=False)
